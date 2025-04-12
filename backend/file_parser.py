#!/usr/bin/env python3
import io
import os
import sys
import json
import zipfile
import shutil
from html import escape
from collections import deque
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from marker.converters.pdf import PdfConverter
from marker.models import create_model_dict
from marker.output import text_from_rendered

# -----------------------------
# Global Directories and Progress Helpers
# -----------------------------
# We will use an internal temporary processing folder
PROCESSING_DIR = os.path.join(os.getcwd(), "processing")
ANALYSIS_DIR = os.path.join(os.getcwd(), "analysis")

# Ensure necessary directories exist
os.makedirs(PROCESSING_DIR, exist_ok=True)
os.makedirs(ANALYSIS_DIR, exist_ok=True)


def load_progress(submission_id):
    progress_file = os.path.join(PROCESSING_DIR, f"progress_{submission_id}.json")
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        return {}


def save_progress(submission_id, progress):
    progress_file = os.path.join(PROCESSING_DIR, f"progress_{submission_id}.json")
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(progress, f, indent=2)


# -----------------------------
# Excel/PDF Conversion Functions (unchanged)
# -----------------------------
DEFAULT_FILL_COLORS = {None, "00000000", "FFFFFFFF"}


def get_cell_info(cell):
    val = cell.value
    color = None
    if cell.fill and cell.fill.fill_type:
        color = cell.fill.start_color.rgb
    return {"value": val, "color": color}


def apply_merged_cells(ws, matrix):
    merged_cell_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        base_cell = matrix[min_row - 1][min_col - 1]
        merged_cell_map[(min_row - 1, min_col - 1)] = {
            "min_row": min_row - 1,
            "min_col": min_col - 1,
            "max_row": max_row - 1,
            "max_col": max_col - 1,
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                if (r, c) != (min_row - 1, min_col - 1):
                    matrix[r][c]["value"] = base_cell["value"]
                    matrix[r][c]["color"] = base_cell["color"]
                    matrix[r][c]["merged_into"] = (min_row - 1, min_col - 1)
    return merged_cell_map


def get_sheet_matrix(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    matrix = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            cell_info = get_cell_info(ws.cell(row=r, column=c))
            cell_info["merged_into"] = None
            row_vals.append(cell_info)
        matrix.append(row_vals)
    merged_cell_map = apply_merged_cells(ws, matrix)
    return matrix, merged_cell_map


def is_nonempty(cell_info):
    val = cell_info["value"]
    return val is not None and str(val).strip() != ""


def is_header_cell(cell_info):
    color = cell_info["color"]
    return color not in DEFAULT_FILL_COLORS


def is_title_candidate(matrix, r, c, merged_cell_map):
    cell_key = (r, c)
    if cell_key in merged_cell_map:
        merged_info = merged_cell_map[cell_key]
        if merged_info["colspan"] > 1 and is_nonempty(matrix[r][c]):
            return True
    if is_nonempty(matrix[r][c]) and is_header_cell(matrix[r][c]):
        return True
    return False


def smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1):
    nrows = len(matrix)
    ncols = len(matrix[0]) if nrows else 0
    visited = [[False] * ncols for _ in range(nrows)]
    components = []
    title_cells = set()
    for r in range(nrows):
        for c in range(ncols):
            if is_title_candidate(matrix, r, c, merged_cell_map):
                title_cells.add((r, c))
                if (r, c) in merged_cell_map:
                    merged_info = merged_cell_map[(r, c)]
                    for mr in range(merged_info["min_row"], merged_info["max_row"] + 1):
                        for mc in range(
                            merged_info["min_col"], merged_info["max_col"] + 1
                        ):
                            title_cells.add((mr, mc))
    for r in range(nrows):
        for c in range(ncols):
            if visited[r][c] or not is_nonempty(matrix[r][c]):
                continue

            q = deque()
            q.append((r, c))
            visited[r][c] = True
            component_cells = {(r, c)}
            min_r, max_r = r, r
            min_c, max_c = c, c
            has_title = (r, c) in title_cells
            while q:
                cr, cc = q.popleft()
                min_r = min(min_r, cr)
                max_r = max(max_r, cr)
                min_c = min(min_c, cc)
                max_c = max(max_c, cc)
                for dr in range(-max_v_gap, max_v_gap + 1):
                    for dc in range(-max_h_gap, max_h_gap + 1):
                        if dr == 0 and dc == 0:
                            continue
                        nr, nc = cr + dr, cc + dc
                        if 0 <= nr < nrows and 0 <= nc < ncols:
                            if not visited[nr][nc] and is_nonempty(matrix[nr][nc]):
                                if abs(dr) > 1 or abs(dc) > 1:
                                    if not is_reasonable_gap(matrix, cr, cc, nr, nc):
                                        continue
                                visited[nr][nc] = True
                                q.append((nr, nc))
                                component_cells.add((nr, nc))
                                if (nr, nc) in title_cells:
                                    has_title = True
            title_row = None
            if not has_title:
                for tr in range(max(0, min_r - 3), min_r):
                    for tc in range(min_c, max_c + 1):
                        if is_title_candidate(matrix, tr, tc, merged_cell_map):
                            title_row = tr
                            for ttc in range(min_c, max_c + 1):
                                if not visited[tr][ttc]:
                                    visited[tr][ttc] = True
                                    component_cells.add((tr, ttc))
                            break
                    if title_row is not None:
                        break
                if title_row is not None:
                    min_r = min(min_r, title_row)
            for fr in range(max_r + 1, min(nrows, max_r + max_v_gap + 1)):
                footer_found = False
                for fc in range(min_c, max_c + 1):
                    if is_nonempty(matrix[fr][fc]):
                        footer_found = True
                        for ffc in range(min_c, max_c + 1):
                            if not visited[fr][ffc]:
                                visited[fr][ffc] = True
                                component_cells.add((fr, ffc))
                        max_r = fr
                        break
                if not footer_found:
                    break
            expanded_cells = set()
            for cell in component_cells:
                r_cell, c_cell = cell
                if matrix[r_cell][c_cell]["merged_into"] is not None:
                    main_r, main_c = matrix[r_cell][c_cell]["merged_into"]
                    expanded_cells.add((main_r, main_c))
                else:
                    expanded_cells.add(cell)
                    if (r_cell, c_cell) in merged_cell_map:
                        merged_info = merged_cell_map[(r_cell, c_cell)]
                        for mr in range(
                            merged_info["min_row"], merged_info["max_row"] + 1
                        ):
                            for mc in range(
                                merged_info["min_col"], merged_info["max_col"] + 1
                            ):
                                expanded_cells.add((mr, mc))
            if expanded_cells:
                min_r = min(r for r, _ in expanded_cells)
                max_r = max(r for r, _ in expanded_cells)
                min_c = min(c for _, c in expanded_cells)
                max_c = max(c for _, c in expanded_cells)
            components.append(
                {
                    "bounds": (min_r, max_r, min_c, max_c),
                    "cells": expanded_cells,
                }
            )
    components.sort(key=lambda comp: len(comp["cells"]), reverse=True)
    final_components = []
    used_cells = set()
    for comp in components:
        if not comp["cells"].intersection(used_cells):
            final_components.append(comp["bounds"])
            used_cells.update(comp["cells"])
    final_components.sort(key=lambda b: (b[0], b[2]))
    return final_components


def is_reasonable_gap(matrix, r1, c1, r2, c2):
    if r1 == r2 or c1 == c2:
        return True
    if abs(r2 - r1) == 2 and c1 == c2:
        middle_r = (r1 + r2) // 2
        return not is_nonempty(matrix[middle_r][c1])
    if abs(c2 - c1) == 2 and r1 == r2:
        middle_c = (c1 + c2) // 2
        return not is_nonempty(matrix[r1][middle_c])
    if abs(r2 - r1) <= 2 and abs(c2 - c1) <= 2:
        return True
    return False


def extract_title(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            cell_key = (r, c)
            if cell_key in merged_cell_map:
                merged_info = merged_cell_map[cell_key]
                span_ratio = (merged_info["max_col"] - merged_info["min_col"] + 1) / (
                    right - left + 1
                )
                if span_ratio >= 0.5 and is_nonempty(matrix[r][c]):
                    return str(matrix[r][c]["value"]).strip(), r < top
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue
        non_empty_cells = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if is_nonempty(matrix[r][c]):
                non_empty_cells.append((r, c))
        if len(non_empty_cells) == 1:
            r_idx, c_idx = non_empty_cells[0]
            return str(matrix[r_idx][c_idx]["value"]).strip(), r_idx < top
    title_parts = []
    for c in range(left, right + 1):
        if c >= len(matrix[0]):
            continue
        if is_nonempty(matrix[top][c]) and matrix[top][c]["merged_into"] is None:
            title_parts.append(str(matrix[top][c]["value"]).strip())
    if title_parts:
        return " ".join(title_parts), False
    return "Untitled Table", False


def dataframe_to_html(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    html_lines = ["<table>"]
    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue
        row_html = ["<tr>"]
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    continue
            cell = matrix[r][c]
            attrs = ""
            if (r, c) in merged_cell_map:
                merged_info = merged_cell_map[(r, c)]
                rowspan = merged_info["rowspan"]
                colspan = merged_info["colspan"]
                effective_rowspan = min(rowspan, bottom - r + 1)
                effective_colspan = min(colspan, right - c + 1)
                if effective_rowspan > 1:
                    attrs += f' rowspan="{effective_rowspan}"'
                if effective_colspan > 1:
                    attrs += f' colspan="{effective_colspan}"'
            cell_content = escape(str(cell["value"])) if is_nonempty(cell) else ""
            row_html.append(f"<td{attrs}>{cell_content}</td>")
        row_html.append("</tr>")
        html_lines.append("".join(row_html))
    html_lines.append("</table>")
    return "\n".join(html_lines)


def dataframe_to_markdown(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    table_grid = []
    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue
        row = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    row.append("")
                    continue
            cell = matrix[r][c]
            cell_content = (
                str(cell["value"]).strip().replace("|", "\\|")
                if is_nonempty(cell)
                else ""
            )
            row.append(cell_content)
        table_grid.append(row)
    if not table_grid:
        return "*Empty table*"
    col_widths = [0] * len(table_grid[0])
    for row in table_grid:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(cell))
    md_lines = []
    header = (
        "| "
        + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(table_grid[0]))
        + " |"
    )
    md_lines.append(header)
    separator = "| " + " | ".join("-" * max(3, width) for width in col_widths) + " |"
    md_lines.append(separator)
    for row_idx, row in enumerate(table_grid):
        if row_idx == 0:
            continue
        md_row = (
            "| "
            + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(row))
            + " |"
        )
        md_lines.append(md_row)
    return "\n".join(md_lines)


def create_markdown_from_workbook_bytes(file_bytes: bytes) -> str:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise Exception(f"Error loading workbook: {e}")
    md_lines = ["# Extracted Excel Subtables", ""]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"## Sheet: {sheet_name}")
        md_lines.append("")
        sheet_md = process_sheet_markdown(ws)
        md_lines.append(sheet_md)
        md_lines.append("")
    return "\n".join(md_lines)


def process_sheet_markdown(ws):
    matrix, merged_cell_map = get_sheet_matrix(ws)
    if not matrix:
        return "*No data found in this sheet.*"
    blocks = smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1)
    md_fragments = []
    for bounds in blocks:
        top, bottom, left, right = bounds
        title, is_external_title = extract_title(matrix, bounds, merged_cell_map)
        md_fragments.append(f"### {title}")
        md_fragments.append("")
        table_top = (
            top + 1 if not is_external_title and title != "Untitled Table" else top
        )
        if table_top <= bottom:
            table_md = dataframe_to_markdown(
                matrix, (table_top, bottom, left, right), merged_cell_map
            )
            md_fragments.append(table_md)
            md_fragments.append("")
        md_fragments.append("---")
        md_fragments.append("")
    return "\n".join(md_fragments)


# -----------------------------
# Processing Functions for Files and Folders
# -----------------------------
def process_excel_file(fpath):
    """Process a single Excel file and return markdown content."""
    try:
        with open(fpath, "rb") as f:
            file_bytes = f.read()
        return create_markdown_from_workbook_bytes(file_bytes)
    except Exception as e:
        return f"Error processing Excel file: {str(e)}"


def process_pdf_file(fpath, output_format="md"):
    """
    Process a PDF file and return its converted contents.

    This function writes the PDF file to a temporary location (since PdfConverter
    requires a file path), calls the converter, and then returns the output
    in Markdown or HTML format.
    """
    try:
        with open(fpath, "rb") as f:
            file_bytes = f.read()
        # Write the PDF to a temporary file because PdfConverter requires a file path.
        with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp.flush()
            tmp_filename = tmp.name

        # Create a PDF converter and run it on the temporary file.
        converter = PdfConverter(artifact_dict=create_model_dict())
        rendered = converter(tmp_filename)
        text, _, images = text_from_rendered(rendered)

        # Cleanup the temporary file.
        os.remove(tmp_filename)

    except Exception as e:
        return f"Error processing PDF file: {str(e)}"

    # For PDF, default to markdown output.
    if output_format.lower() == "md":
        result = text
    else:
        # If HTML is requested, you could convert the markdown output to HTML.
        # For now, we simply return the same text.
        result = text
    return result


def process_zip_file(fpath, temp_dir):
    """Extract a ZIP archive and return a list of paths to extracted files."""
    extracted = []
    with zipfile.ZipFile(fpath, "r") as zip_ref:
        zip_ref.extractall(temp_dir)
        extracted = [os.path.join(temp_dir, name) for name in zip_ref.namelist()]
    return extracted


def process_file(fpath):
    """
    Determine file type and process Excel or PDF files.
    Returns a tuple of (filename, markdown_content) if processed, or None otherwise.
    """
    ext = os.path.splitext(fpath)[1].lower()
    if ext == ".xlsx":
        return os.path.basename(fpath), process_excel_file(fpath)
    elif ext == ".pdf":
        return os.path.basename(fpath), process_pdf_file(fpath)
    else:
        return None


def process_folder(input_folder, output_folder):
    """
    For each file in the input folder:
      - If the file is a ZIP archive, extract and process its contents.
      - For Excel (.xlsx) and PDF (.pdf) files, convert to markdown.
      - For unsupported files, copy them to the output folder.
    """
    if not os.path.isdir(input_folder):
        print(f"Error: {input_folder} is not a directory.")
        sys.exit(1)
    os.makedirs(output_folder, exist_ok=True)

    # List all items in the input folder (non-recursive)
    for item in os.listdir(input_folder):
        full_path = os.path.join(input_folder, item)
        if os.path.isfile(full_path):
            ext = os.path.splitext(full_path)[1].lower()
            # Create a temporary processing folder for ZIP extraction
            temp_processing_folder = os.path.join(PROCESSING_DIR, "temp")
            os.makedirs(temp_processing_folder, exist_ok=True)

            if ext == ".zip" and zipfile.is_zipfile(full_path):
                print(f"Extracting ZIP archive: {item}")
                extracted_files = process_zip_file(full_path, temp_processing_folder)
                for extracted in extracted_files:
                    result = process_file(extracted)
                    if result:
                        out_filename, markdown = result
                        output_path = os.path.join(output_folder, out_filename + ".md")
                        with open(output_path, "w", encoding="utf-8") as outf:
                            outf.write(markdown)
                        print(
                            f"Processed {out_filename} from ZIP and wrote to {output_path}"
                        )
                    else:
                        # Unsupported file: copy it to output
                        dest_path = os.path.join(
                            output_folder, os.path.basename(extracted)
                        )
                        shutil.copy(extracted, dest_path)
                        print(
                            f"Copied unsupported file from ZIP: {os.path.basename(extracted)} to {dest_path}"
                        )
                # Clean up temp folder after processing zip
                shutil.rmtree(temp_processing_folder)
            elif ext in [".xlsx", ".pdf"]:
                result = process_file(full_path)
                if result:
                    out_filename, markdown = result
                    output_path = os.path.join(output_folder, out_filename + ".md")
                    with open(output_path, "w", encoding="utf-8") as outf:
                        outf.write(markdown)
                    print(f"Processed {item} and wrote to {output_path}")
            else:
                # Unsupported file: copy it to output folder
                dest_path = os.path.join(output_folder, item)
                shutil.copy(full_path, dest_path)
                print(f"Copied unsupported file: {item} to {dest_path}")


# -----------------------------
# Main Entry Point
# -----------------------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python process_folder.py <input_folder> <output_folder>")
        sys.exit(1)
    input_folder = sys.argv[1]
    output_folder = sys.argv[2]
    process_folder(input_folder, output_folder)

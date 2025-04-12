import io
import os
import time
import json
import shutil
import zipfile
import asyncio
from uuid import uuid4
from datetime import datetime
from threading import Lock
from typing import List

from fastapi import (
    FastAPI,
    UploadFile,
    File,
    HTTPException,
    Path,
    BackgroundTasks,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from html import escape
from collections import deque
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook

# -----------------------------
# Directory Setup
# -----------------------------
UPLOAD_DIR = "uploads"
PROCESSING_DIR = "processing"
PROCESSED_DIR = "processed"
ANALYSIS_DIR = "analysis"
PROGRESS_DIR = "progress"  # Persist progress per submission

for folder in [UPLOAD_DIR, PROCESSING_DIR, PROCESSED_DIR, ANALYSIS_DIR, PROGRESS_DIR]:
    os.makedirs(folder, exist_ok=True)

# Lock for progress file writes
progress_lock = Lock()


# -----------------------------
# Utility: Persistent progress
# -----------------------------
def get_progress_path(submission_id: str) -> str:
    return os.path.join(PROGRESS_DIR, f"progress_{submission_id}.json")


def load_progress(submission_id: str) -> dict:
    progress_path = get_progress_path(submission_id)
    if os.path.exists(progress_path):
        with open(progress_path, "r") as f:
            return json.load(f)
    return {}


def save_progress(submission_id: str, data: dict):
    progress_path = get_progress_path(submission_id)
    with progress_lock:
        with open(progress_path, "w") as f:
            json.dump(data, f, indent=2)


# -----------------------------
# Excel/PDF Conversion Functions
# -----------------------------
DEFAULT_FILL_COLORS = {None, "00000000", "FFFFFFFF"}


def get_cell_info(cell):
    val = cell.value
    color = None
    if cell.fill and cell.fill.fill_type:
        color = cell.fill.start_color.rgb
    return {"value": val, "color": color}


def apply_merged_cells(ws, matrix):
    merged_cell_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        base_cell = matrix[min_row - 1][min_col - 1]
        merged_cell_map[(min_row - 1, min_col - 1)] = {
            "min_row": min_row - 1,
            "min_col": min_col - 1,
            "max_row": max_row - 1,
            "max_col": max_col - 1,
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                if (r, c) != (min_row - 1, min_col - 1):
                    matrix[r][c]["value"] = base_cell["value"]
                    matrix[r][c]["color"] = base_cell["color"]
                    matrix[r][c]["merged_into"] = (min_row - 1, min_col - 1)
    return merged_cell_map


def get_sheet_matrix(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    matrix = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            cell_info = get_cell_info(ws.cell(row=r, column=c))
            cell_info["merged_into"] = None
            row_vals.append(cell_info)
        matrix.append(row_vals)
    merged_cell_map = apply_merged_cells(ws, matrix)
    return matrix, merged_cell_map


def is_nonempty(cell_info):
    val = cell_info["value"]
    return val is not None and str(val).strip() != ""


def is_header_cell(cell_info):
    color = cell_info["color"]
    return color not in DEFAULT_FILL_COLORS


def is_title_candidate(matrix, r, c, merged_cell_map):
    cell_key = (r, c)
    if cell_key in merged_cell_map:
        merged_info = merged_cell_map[cell_key]
        if merged_info["colspan"] > 1 and is_nonempty(matrix[r][c]):
            return True
    if is_nonempty(matrix[r][c]) and is_header_cell(matrix[r][c]):
        return True
    return False


def smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1):
    nrows = len(matrix)
    ncols = len(matrix[0]) if nrows else 0
    visited = [[False] * ncols for _ in range(nrows)]
    components = []
    title_cells = set()
    for r in range(nrows):
        for c in range(ncols):
            if is_title_candidate(matrix, r, c, merged_cell_map):
                title_cells.add((r, c))
                if (r, c) in merged_cell_map:
                    merged_info = merged_cell_map[(r, c)]
                    for mr in range(merged_info["min_row"], merged_info["max_row"] + 1):
                        for mc in range(
                            merged_info["min_col"], merged_info["max_col"] + 1
                        ):
                            title_cells.add((mr, mc))
    for r in range(nrows):
        for c in range(ncols):
            if visited[r][c] or not is_nonempty(matrix[r][c]):
                continue

            q = deque()
            q.append((r, c))
            visited[r][c] = True
            component_cells = {(r, c)}
            min_r, max_r = r, r
            min_c, max_c = c, c
            has_title = (r, c) in title_cells
            while q:
                cr, cc = q.popleft()
                min_r = min(min_r, cr)
                max_r = max(max_r, cr)
                min_c = min(min_c, cc)
                max_c = max(max_c, cc)
                for dr in range(-max_v_gap, max_v_gap + 1):
                    for dc in range(-max_h_gap, max_h_gap + 1):
                        if dr == 0 and dc == 0:
                            continue
                        nr, nc = cr + dr, cc + dc
                        if 0 <= nr < nrows and 0 <= nc < ncols:
                            if not visited[nr][nc] and is_nonempty(matrix[nr][nc]):
                                if abs(dr) > 1 or abs(dc) > 1:
                                    if not is_reasonable_gap(matrix, cr, cc, nr, nc):
                                        continue
                                visited[nr][nc] = True
                                q.append((nr, nc))
                                component_cells.add((nr, nc))
                                if (nr, nc) in title_cells:
                                    has_title = True
            title_row = None
            if not has_title:
                for tr in range(max(0, min_r - 3), min_r):
                    for tc in range(min_c, max_c + 1):
                        if is_title_candidate(matrix, tr, tc, merged_cell_map):
                            title_row = tr
                            for ttc in range(min_c, max_c + 1):
                                if not visited[tr][ttc]:
                                    visited[tr][ttc] = True
                                    component_cells.add((tr, ttc))
                            break
                    if title_row is not None:
                        break
                if title_row is not None:
                    min_r = min(min_r, title_row)
            for fr in range(max_r + 1, min(nrows, max_r + max_v_gap + 1)):
                footer_found = False
                for fc in range(min_c, max_c + 1):
                    if is_nonempty(matrix[fr][fc]):
                        footer_found = True
                        for ffc in range(min_c, max_c + 1):
                            if not visited[fr][ffc]:
                                visited[fr][ffc] = True
                                component_cells.add((fr, ffc))
                        max_r = fr
                        break
                if not footer_found:
                    break
            expanded_cells = set()
            for cell in component_cells:
                r_cell, c_cell = cell
                if matrix[r_cell][c_cell]["merged_into"] is not None:
                    main_r, main_c = matrix[r_cell][c_cell]["merged_into"]
                    expanded_cells.add((main_r, main_c))
                else:
                    expanded_cells.add(cell)
                    if (r_cell, c_cell) in merged_cell_map:
                        merged_info = merged_cell_map[(r_cell, c_cell)]
                        for mr in range(
                            merged_info["min_row"], merged_info["max_row"] + 1
                        ):
                            for mc in range(
                                merged_info["min_col"], merged_info["max_col"] + 1
                            ):
                                expanded_cells.add((mr, mc))
            if expanded_cells:
                min_r = min(r for r, _ in expanded_cells)
                max_r = max(r for r, _ in expanded_cells)
                min_c = min(c for _, c in expanded_cells)
                max_c = max(c for _, c in expanded_cells)
            components.append(
                {
                    "bounds": (min_r, max_r, min_c, max_c),
                    "cells": expanded_cells,
                }
            )
    components.sort(key=lambda comp: len(comp["cells"]), reverse=True)
    final_components = []
    used_cells = set()
    for comp in components:
        if not comp["cells"].intersection(used_cells):
            final_components.append(comp["bounds"])
            used_cells.update(comp["cells"])
    final_components.sort(key=lambda b: (b[0], b[2]))
    return final_components


def is_reasonable_gap(matrix, r1, c1, r2, c2):
    if r1 == r2 or c1 == c2:
        return True
    if abs(r2 - r1) == 2 and c1 == c2:
        middle_r = (r1 + r2) // 2
        return not is_nonempty(matrix[middle_r][c1])
    if abs(c2 - c1) == 2 and r1 == r2:
        middle_c = (c1 + c2) // 2
        return not is_nonempty(matrix[r1][middle_c])
    if abs(r2 - r1) <= 2 and abs(c2 - c1) <= 2:
        return True
    return False


def extract_title(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            cell_key = (r, c)
            if cell_key in merged_cell_map:
                merged_info = merged_cell_map[cell_key]
                span_ratio = (merged_info["max_col"] - merged_info["min_col"] + 1) / (
                    right - left + 1
                )
                if span_ratio >= 0.5 and is_nonempty(matrix[r][c]):
                    return str(matrix[r][c]["value"]).strip(), r < top
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue
        non_empty_cells = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if is_nonempty(matrix[r][c]):
                non_empty_cells.append((r, c))
        if len(non_empty_cells) == 1:
            r_idx, c_idx = non_empty_cells[0]
            return str(matrix[r_idx][c_idx]["value"]).strip(), r_idx < top
    title_parts = []
    for c in range(left, right + 1):
        if c >= len(matrix[0]):
            continue
        if is_nonempty(matrix[top][c]) and matrix[top][c]["merged_into"] is None:
            title_parts.append(str(matrix[top][c]["value"]).strip())
    if title_parts:
        return " ".join(title_parts), False
    return "Untitled Table", False


def dataframe_to_html(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    html_lines = ["<table>"]
    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue
        row_html = ["<tr>"]
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    continue
            cell = matrix[r][c]
            attrs = ""
            if (r, c) in merged_cell_map:
                merged_info = merged_cell_map[(r, c)]
                rowspan = merged_info["rowspan"]
                colspan = merged_info["colspan"]
                effective_rowspan = min(rowspan, bottom - r + 1)
                effective_colspan = min(colspan, right - c + 1)
                if effective_rowspan > 1:
                    attrs += f' rowspan="{effective_rowspan}"'
                if effective_colspan > 1:
                    attrs += f' colspan="{effective_colspan}"'
            cell_content = escape(str(cell["value"])) if is_nonempty(cell) else ""
            row_html.append(f"<td{attrs}>{cell_content}</td>")
        row_html.append("</tr>")
        html_lines.append("".join(row_html))
    html_lines.append("</table>")
    return "\n".join(html_lines)


def dataframe_to_markdown(matrix, bounds, merged_cell_map):
    top, bottom, left, right = bounds
    table_grid = []
    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue
        row = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    row.append("")
                    continue
            cell = matrix[r][c]
            cell_content = (
                str(cell["value"]).strip().replace("|", "\\|")
                if is_nonempty(cell)
                else ""
            )
            row.append(cell_content)
        table_grid.append(row)
    if not table_grid:
        return "*Empty table*"
    col_widths = [0] * len(table_grid[0])
    for row in table_grid:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(cell))
    md_lines = []
    header = (
        "| "
        + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(table_grid[0]))
        + " |"
    )
    md_lines.append(header)
    separator = "| " + " | ".join("-" * max(3, width) for width in col_widths) + " |"
    md_lines.append(separator)
    for row_idx, row in enumerate(table_grid):
        if row_idx == 0:
            continue
        md_row = (
            "| "
            + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(row))
            + " |"
        )
        md_lines.append(md_row)
    return "\n".join(md_lines)


def create_html_from_workbook_bytes(file_bytes: bytes) -> str:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error loading workbook: {e}")
    html_lines = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        '    <meta charset="UTF-8">',
        "    <title>Extracted Excel Subtables</title>",
        "</head>",
        "<body>",
        "<h1>Extracted Subtables</h1>",
    ]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_lines.append("<section>")
        html_lines.append(f"<h2>Sheet: {escape(sheet_name)}</h2>")
        sheet_html = process_sheet_html(ws)
        html_lines.append(sheet_html)
        html_lines.append("</section>")
    html_lines.append("</body>")
    html_lines.append("</html>")
    return "\n".join(html_lines)


def create_markdown_from_workbook_bytes(file_bytes: bytes) -> str:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error loading workbook: {e}")
    md_lines = ["# Extracted Excel Subtables", ""]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"## Sheet: {sheet_name}")
        md_lines.append("")
        sheet_md = process_sheet_markdown(ws)
        md_lines.append(sheet_md)
        md_lines.append("")
    return "\n".join(md_lines)


def process_sheet_html(ws):
    matrix, merged_cell_map = get_sheet_matrix(ws)
    if not matrix:
        return "<p>No data found in this sheet.</p>"
    blocks = smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1)
    html_fragments = []
    for bounds in blocks:
        top, bottom, left, right = bounds
        title, is_external_title = extract_title(matrix, bounds, merged_cell_map)
        html_fragments.append(f"<h3>{escape(title)}</h3>")
        table_top = (
            top + 1 if not is_external_title and title != "Untitled Table" else top
        )
        if table_top <= bottom:
            table_html = dataframe_to_html(
                matrix, (table_top, bottom, left, right), merged_cell_map
            )
            html_fragments.append(table_html)
        html_fragments.append("<hr>")
    return "\n".join(html_fragments)


def process_sheet_markdown(ws):
    matrix, merged_cell_map = get_sheet_matrix(ws)
    if not matrix:
        return "*No data found in this sheet.*"
    blocks = smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1)
    md_fragments = []
    for bounds in blocks:
        top, bottom, left, right = bounds
        title, is_external_title = extract_title(matrix, bounds, merged_cell_map)
        md_fragments.append(f"### {title}")
        md_fragments.append("")
        table_top = (
            top + 1 if not is_external_title and title != "Untitled Table" else top
        )
        if table_top <= bottom:
            table_md = dataframe_to_markdown(
                matrix, (table_top, bottom, left, right), merged_cell_map
            )
            md_fragments.append(table_md)
            md_fragments.append("")
        md_fragments.append("---")
        md_fragments.append("")
    return "\n".join(md_fragments)


# -----------------------------
# FastAPI Application Initialization
# -----------------------------
app = FastAPI(title="File Processing API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve processed markdown files as static content (for testing)
app.mount("/results", StaticFiles(directory=PROCESSED_DIR), name="results")


# -----------------------------
# Submission and File Processing Endpoints
# -----------------------------
@app.post("/submissions")
async def create_submission(
    background_tasks: BackgroundTasks, file: UploadFile = File(...)
):
    # Save uploaded file to disk
    submission_id = str(uuid4())
    submission_folder = os.path.join(UPLOAD_DIR, submission_id)
    os.makedirs(submission_folder, exist_ok=True)
    file_path = os.path.join(submission_folder, file.filename)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    # Create submission record with initial progress saved to disk
    submission = {
        "id": submission_id,
        "status": "pending",
        "createdAt": datetime.utcnow().isoformat(),
        "files": [file.filename],
        "progress": "File uploaded, waiting for processing",
    }
    save_progress(submission_id, submission)
    # Start background processing task
    background_tasks.add_task(process_submission, submission_id, file_path)
    return submission


@app.get("/submissions")
async def list_submissions():
    # List all submissions by scanning PROGRESS_DIR
    submissions = []
    for fname in os.listdir(PROGRESS_DIR):
        if fname.startswith("progress_") and fname.endswith(".json"):
            submission_id = fname[len("progress_") : -len(".json")]
            submissions.append(load_progress(submission_id))
    return submissions


@app.get("/submissions/{submission_id}")
async def get_submission_status(submission_id: str = Path(...)):
    progress = load_progress(submission_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Submission not found")
    return progress


@app.post("/submissions/{submission_id}/files")
async def upload_additional_files(
    submission_id: str = Path(...),
    background_tasks: BackgroundTasks = None,
    files: List[UploadFile] = File(...),
):
    progress = load_progress(submission_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Submission not found")
    submission_folder = os.path.join(UPLOAD_DIR, submission_id)
    os.makedirs(submission_folder, exist_ok=True)
    for file in files:
        file_path = os.path.join(submission_folder, file.filename)
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        progress["files"].append(file.filename)
        background_tasks.add_task(process_submission, submission_id, file_path)
    progress["progress"] = "Additional files uploaded, processing started"
    save_progress(submission_id, progress)
    return progress


# -----------------------------
# WebSocket Endpoint for Live Status Updates
# -----------------------------
@app.websocket("/ws/{submission_id}")
async def websocket_endpoint(websocket: WebSocket, submission_id: str):
    await websocket.accept()
    try:
        while True:
            progress = load_progress(submission_id)
            await websocket.send_json(progress)
            await asyncio.sleep(2)
    except WebSocketDisconnect:
        print(f"WebSocket disconnected for submission {submission_id}")


# -----------------------------
# Background Processing Function
# -----------------------------
def process_submission(submission_id: str, file_path: str):
    """
    Process an uploaded file:
      - If compressed (zip), extract its contents.
      - Search for Excel (.xlsx) or PDF (.pdf) files.
      - For each, convert them to markdown.
      - Save the generated markdown file to the processed folder.
      - Move the original processed file to the analysis folder.
      - Update persistent progress along the way.
    """
    progress = load_progress(submission_id)
    progress["progress"] = "Processing started"
    save_progress(submission_id, progress)

    processing_subdir = os.path.join(PROCESSING_DIR, submission_id)
    os.makedirs(processing_subdir, exist_ok=True)

    extracted_files = []
    if zipfile.is_zipfile(file_path):
        progress["progress"] = "Extracting ZIP archive"
        save_progress(submission_id, progress)
        with zipfile.ZipFile(file_path, "r") as zip_ref:
            zip_ref.extractall(processing_subdir)
            extracted_files = [
                os.path.join(processing_subdir, name) for name in zip_ref.namelist()
            ]
    else:
        dest_path = os.path.join(processing_subdir, os.path.basename(file_path))
        shutil.copy(file_path, dest_path)
        extracted_files = [dest_path]

    results = {}
    for fpath in extracted_files:
        ext = os.path.splitext(fpath)[1].lower()
        result_md = ""
        if ext == ".xlsx":
            try:
                with open(fpath, "rb") as f:
                    file_bytes = f.read()
                result_md = create_markdown_from_workbook_bytes(file_bytes)
            except Exception as e:
                result_md = f"Error processing Excel file: {str(e)}"
        elif ext == ".pdf":
            try:
                with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(open(fpath, "rb").read())
                    tmp.flush()
                    tmp_filename = tmp.name
                from marker.converters.pdf import PdfConverter
                from marker.models import create_model_dict
                from marker.output import text_from_rendered

                converter = PdfConverter(artifact_dict=create_model_dict())
                rendered = converter(tmp_filename)
                text, _, images = text_from_rendered(rendered)
                os.remove(tmp_filename)
                result_md = text
            except Exception as e:
                result_md = f"Error processing PDF file: {str(e)}"
        else:
            continue
        if result_md:
            submission_processed_dir = os.path.join(PROCESSED_DIR, submission_id)
            os.makedirs(submission_processed_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(fpath))[0]
            out_file = os.path.join(submission_processed_dir, f"{base_name}.md")
            with open(out_file, "w", encoding="utf-8") as outf:
                outf.write(result_md)
            results[os.path.basename(fpath)] = out_file
        progress["progress"] = f"Processed {os.path.basename(fpath)}"
        save_progress(submission_id, progress)
        time.sleep(1)

    analysis_subdir = os.path.join(ANALYSIS_DIR, submission_id)
    os.makedirs(analysis_subdir, exist_ok=True)
    for f in extracted_files:
        shutil.move(f, analysis_subdir)

    progress["status"] = "complete"
    progress["progress"] = "All files processed"
    progress["results"] = results
    progress["processedAt"] = datetime.utcnow().isoformat()
    save_progress(submission_id, progress)


# -----------------------------
# Endpoints for File Operations
# -----------------------------
@app.get("/files/{submission_id}")
async def get_submission_files(submission_id: str = Path(...)):
    progress = load_progress(submission_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Submission not found")
    return {"files": progress.get("files", []), "results": progress.get("results", {})}


@app.get("/files/{submission_id}/{file_name}/content")
async def get_file_content(submission_id: str = Path(...), file_name: str = Path(...)):
    # Look in the processed folder for a markdown file matching the base name.
    submission_processed_dir = os.path.join(PROCESSED_DIR, submission_id)
    base_name = os.path.splitext(file_name)[0]
    md_file = os.path.join(submission_processed_dir, f"{base_name}.md")
    if not os.path.exists(md_file):
        raise HTTPException(status_code=404, detail="Processed file not found")
    with open(md_file, "r", encoding="utf-8") as f:
        content = f.read()
    return {"text": content}


@app.get("/files/{submission_id}/{file_name}/download")
async def download_file(submission_id: str = Path(...), file_name: str = Path(...)):
    # Try to locate the original file in the analysis folder first.
    analysis_subdir = os.path.join(ANALYSIS_DIR, submission_id)
    file_path = os.path.join(analysis_subdir, file_name)
    if not os.path.exists(file_path):
        # Fall back to looking in the upload folder.
        upload_subdir = os.path.join(UPLOAD_DIR, submission_id)
        file_path = os.path.join(upload_subdir, file_name)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        file_path, media_type="application/octet-stream", filename=file_name
    )


@app.delete("/files/{submission_id}/{file_name}")
async def delete_file(submission_id: str = Path(...), file_name: str = Path(...)):
    # Remove file from analysis folder if exists.
    analysis_subdir = os.path.join(ANALYSIS_DIR, submission_id)
    file_path = os.path.join(analysis_subdir, file_name)
    if os.path.exists(file_path):
        os.remove(file_path)
    # Also remove from processed folder (if markdown exists)
    submission_processed_dir = os.path.join(PROCESSED_DIR, submission_id)
    base_name = os.path.splitext(file_name)[0]
    md_file = os.path.join(submission_processed_dir, f"{base_name}.md")
    if os.path.exists(md_file):
        os.remove(md_file)
    progress = load_progress(submission_id)
    if progress and "files" in progress:
        progress["files"] = [f for f in progress["files"] if f != file_name]
        save_progress(submission_id, progress)
    return {"success": True, "message": "File deleted successfully"}


@app.post("/files/{submission_id}/compare-contracts")
async def compare_contracts(submission_id: str = Path(...), payload: dict = None):
    if not payload or "previousFile" not in payload or "currentFile" not in payload:
        raise HTTPException(
            status_code=400, detail="Payload must include previousFile and currentFile"
        )
    # Simulate contract comparison.
    comparison = {
        "clauses": [
            {
                "id": "clause-1",
                "name": "Coverage Territory",
                "previousText": "This policy covers properties located in Florida, excluding Monroe County.",
                "currentText": "Now covers Florida excluding Monroe County and coastal areas within 5 miles.",
                "significance": "high",
                "changes": [
                    {"type": "modification", "text": "Added coastal area exclusion."}
                ],
            }
        ]
    }
    return comparison


@app.post("/files/{submission_id}/analyze-claims")
async def analyze_claims(submission_id: str = Path(...), payload: dict = None):
    if not payload or "fileName" not in payload:
        raise HTTPException(status_code=400, detail="Payload must include fileName")
    analysis = {
        "analysis": f"Simulated analysis of claims data from {payload['fileName']}."
    }
    return analysis


@app.post("/files/{submission_id}/analyze-exposure")
async def analyze_exposure(submission_id: str = Path(...), payload: dict = None):
    if not payload or "fileName" not in payload:
        raise HTTPException(status_code=400, detail="Payload must include fileName")
    analysis = {
        "analysis": f"Simulated analysis of exposure data from {payload['fileName']}."
    }
    return analysis


@app.post("/files/{submission_id}/extract-insights")
async def extract_insights(submission_id: str = Path(...), payload: dict = None):
    if not payload or "fileNames" not in payload:
        raise HTTPException(
            status_code=400, detail="Payload must include a list of fileNames"
        )
    insights = {
        "insights": f"Simulated insights extracted from files: {', '.join(payload['fileNames'])}."
    }
    return insights


# -----------------------------
# Dashboard Endpoint
# -----------------------------
@app.get("/dashboard/{submission_id}")
async def get_dashboard_data(submission_id: str = Path(...)):
    progress = load_progress(submission_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Submission not found")
    # Simulated dashboard data combining progress and some dummy values.
    dashboard = {
        "contractRiskData": {"riskScore": 75},
        "layerData": {"layerCount": 3},
        "hurricaneData": {"peakWind": 120},
        "claimsData": {"totalClaims": 5},
        "exposureData": {"exposureValue": 1000000},
        "economicData": {"lossEstimate": 250000},
        "climateRiskData": {"climateIndex": 0.85},
        "contractChanges": progress.get("results", {}),
        "keyInsights": {"summary": "Processed successfully"},
        "underwriterResponse": progress.get("underwriterResponse", {}),
    }
    return dashboard


@app.post("/submissions/{submission_id}/response")
async def submit_underwriter_response(
    submission_id: str = Path(...), response_data: dict = None
):
    progress = load_progress(submission_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Submission not found")
    progress["underwriterResponse"] = response_data or {}
    save_progress(submission_id, progress)
    return {"success": True, "message": "Underwriter response submitted successfully"}


# -----------------------------
# Main Execution
# -----------------------------
if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app)

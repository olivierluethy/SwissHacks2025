import io
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import Response
from html import escape
from collections import deque
from openpyxl import load_workbook

app = FastAPI(title="Excel Subtable Extractor")

# A list of RGB values considered as "default" (non‐header).
DEFAULT_FILL_COLORS = {None, "00000000", "FFFFFFFF"}


def get_cell_info(cell):
    """
    Return a dictionary with the cell's value and its fill color (RGB string if available).
    """
    val = cell.value
    color = None
    if cell.fill and cell.fill.fill_type:
        color = cell.fill.start_color.rgb
    return {"value": val, "color": color}


def apply_merged_cells(ws, matrix):
    """
    Propagate values and colors from the top-left cell of merged areas.
    Also return a dictionary mapping cell coordinates to merged cell information.
    """
    merged_cell_map = {}  # Maps (row, col) to merged range bounds

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        base_cell = matrix[min_row - 1][min_col - 1]

        # Store the merged range information
        merged_cell_map[(min_row - 1, min_col - 1)] = {
            "min_row": min_row - 1,
            "min_col": min_col - 1,
            "max_row": max_row - 1,
            "max_col": max_col - 1,
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }

        # Propagate values to all cells in the merged range
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                if (r, c) != (min_row - 1, min_col - 1):
                    matrix[r][c]["value"] = base_cell["value"]
                    matrix[r][c]["color"] = base_cell["color"]
                    # Mark other cells in the merged range
                    matrix[r][c]["merged_into"] = (min_row - 1, min_col - 1)

    return merged_cell_map


def get_sheet_matrix(ws):
    """
    Read the worksheet into a 2D matrix of dictionaries.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    matrix = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            cell_info = get_cell_info(ws.cell(row=r, column=c))
            cell_info["merged_into"] = (
                None  # Will be set in apply_merged_cells if applicable
            )
            row_vals.append(cell_info)
        matrix.append(row_vals)

    merged_cell_map = apply_merged_cells(ws, matrix)
    return matrix, merged_cell_map


def is_nonempty(cell_info):
    """
    Return True if the cell is nonempty.
    """
    val = cell_info["value"]
    return val is not None and str(val).strip() != ""


def is_header_cell(cell_info):
    """
    A cell qualifies as a header if its fill color is not a default color.
    """
    color = cell_info["color"]
    return color not in DEFAULT_FILL_COLORS


def is_title_candidate(matrix, r, c, merged_cell_map):
    """
    Check if a cell is likely to be a title based on:
    1. It's a merged cell that spans multiple columns
    2. It's positioned at the top of a data region
    3. It has content
    """
    cell_key = (r, c)
    if cell_key in merged_cell_map:
        merged_info = merged_cell_map[cell_key]
        # If it spans multiple columns and has content
        if merged_info["colspan"] > 1 and is_nonempty(matrix[r][c]):
            return True

    # Check if it's a standalone cell with content that might be a title
    if is_nonempty(matrix[r][c]) and is_header_cell(matrix[r][c]):
        return True

    return False


def smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1):
    """
    Use BFS to detect connected components of nonempty cells.
    Allows for small gaps (max_gap) to connect nearby cells.
    Ensures each cell is only used in one component (exhaustive search).
    """
    nrows = len(matrix)
    ncols = len(matrix[0]) if nrows else 0
    visited = [[False] * ncols for _ in range(nrows)]
    components = []

    # First pass: identify potential title cells
    title_cells = set()
    for r in range(nrows):
        for c in range(ncols):
            if is_title_candidate(matrix, r, c, merged_cell_map):
                title_cells.add((r, c))
                # Mark the entire merged region as title cells
                if (r, c) in merged_cell_map:
                    merged_info = merged_cell_map[(r, c)]
                    for mr in range(merged_info["min_row"], merged_info["max_row"] + 1):
                        for mc in range(
                            merged_info["min_col"], merged_info["max_col"] + 1
                        ):
                            title_cells.add((mr, mc))

    # Process cells in order (top-to-bottom, left-to-right)
    for r in range(nrows):
        for c in range(ncols):
            # Skip if already visited or empty
            if visited[r][c] or not is_nonempty(matrix[r][c]):
                continue

            # Start a new component
            q = deque()
            q.append((r, c))
            visited[r][c] = True
            component_cells = set([(r, c)])
            min_r, max_r = r, r
            min_c, max_c = c, c

            # Track if this component has a title
            has_title = (r, c) in title_cells

            while q:
                cr, cc = q.popleft()
                min_r = min(min_r, cr)
                max_r = max(max_r, cr)
                min_c = min(min_c, cc)
                max_c = max(max_c, cc)

                # Check all directions including diagonals
                for dr in range(-max_v_gap, max_v_gap + 1):
                    for dc in range(-max_h_gap, max_h_gap + 1):
                        if dr == 0 and dc == 0:
                            continue

                        nr, nc = cr + dr, cc + dc
                        if 0 <= nr < nrows and 0 <= nc < ncols:
                            if not visited[nr][nc] and is_nonempty(matrix[nr][nc]):
                                # For gaps > 1, check if the gap is reasonable
                                if abs(dr) > 1 or abs(dc) > 1:
                                    # Only allow gaps if there's content similarity or alignment
                                    if not is_reasonable_gap(matrix, cr, cc, nr, nc):
                                        continue

                                visited[nr][nc] = True
                                q.append((nr, nc))
                                component_cells.add((nr, nc))

                                # Check if we found a title cell
                                if (nr, nc) in title_cells:
                                    has_title = True

            # Look for potential titles above the component
            title_row = None
            if not has_title:
                for tr in range(max(0, min_r - 3), min_r):
                    for tc in range(min_c, max_c + 1):
                        if 0 <= tr < nrows and 0 <= tc < ncols:
                            if is_title_candidate(matrix, tr, tc, merged_cell_map):
                                title_row = tr
                                # Add title cells to the component
                                for ttc in range(min_c, max_c + 1):
                                    if 0 <= ttc < ncols and not visited[tr][ttc]:
                                        visited[tr][ttc] = True
                                        component_cells.add((tr, ttc))
                                break
                    if title_row is not None:
                        break

                if title_row is not None:
                    min_r = min(min_r, title_row)

            # Look for footers below the component
            for fr in range(max_r + 1, min(nrows, max_r + max_v_gap + 1)):
                footer_found = False
                for fc in range(min_c, max_c + 1):
                    if 0 <= fc < ncols and is_nonempty(matrix[fr][fc]):
                        footer_found = True
                        # Add footer cells to the component
                        for ffc in range(min_c, max_c + 1):
                            if 0 <= ffc < ncols and not visited[fr][ffc]:
                                visited[fr][ffc] = True
                                component_cells.add((fr, ffc))
                        max_r = fr
                        break
                if not footer_found:
                    break

            # Expand to include all cells in merged regions
            expanded_cells = set()
            for cell in component_cells:
                r_cell, c_cell = cell
                if matrix[r_cell][c_cell]["merged_into"] is not None:
                    # This cell is part of a merged region, add the main cell
                    main_r, main_c = matrix[r_cell][c_cell]["merged_into"]
                    expanded_cells.add((main_r, main_c))
                else:
                    expanded_cells.add(cell)
                    # If this is a main merged cell, add all cells in the merged region
                    if (r_cell, c_cell) in merged_cell_map:
                        merged_info = merged_cell_map[(r_cell, c_cell)]
                        for mr in range(
                            merged_info["min_row"], merged_info["max_row"] + 1
                        ):
                            for mc in range(
                                merged_info["min_col"], merged_info["max_col"] + 1
                            ):
                                expanded_cells.add((mr, mc))
                                visited[mr][mc] = True

            # Recalculate bounds based on expanded cells
            if expanded_cells:
                min_r = min(r for r, _ in expanded_cells)
                max_r = max(r for r, _ in expanded_cells)
                min_c = min(c for _, c in expanded_cells)
                max_c = max(c for _, c in expanded_cells)

            components.append(
                {"bounds": (min_r, max_r, min_c, max_c), "cells": expanded_cells}
            )

    # Sort components by size (largest first) to prioritize larger tables
    components.sort(key=lambda comp: len(comp["cells"]), reverse=True)

    # Filter out components that overlap with already selected components
    final_components = []
    used_cells = set()

    for comp in components:
        # Check if this component uses any already used cells
        if not comp["cells"].intersection(used_cells):
            final_components.append(comp["bounds"])
            used_cells.update(comp["cells"])

    # Sort final components by position (top-to-bottom, left-to-right)
    final_components.sort(key=lambda b: (b[0], b[2]))

    return final_components


def is_reasonable_gap(matrix, r1, c1, r2, c2):
    """
    Determine if a gap between cells is reasonable based on content similarity,
    alignment, or other heuristics.
    """
    # Check if cells are in the same row or column (alignment)
    if r1 == r2 or c1 == c2:
        return True

    # Check if there's a pattern of empty cells between them
    if abs(r2 - r1) == 2 and c1 == c2:
        # Check if the cell between is empty (common in some reports)
        middle_r = (r1 + r2) // 2
        return not is_nonempty(matrix[middle_r][c1])

    if abs(c2 - c1) == 2 and r1 == r2:
        # Check if the cell between is empty
        middle_c = (c1 + c2) // 2
        return not is_nonempty(matrix[r1][middle_c])

    # For diagonal gaps, be more conservative
    if abs(r2 - r1) <= 2 and abs(c2 - c1) <= 2:
        return True

    return False


def extract_title(matrix, bounds, merged_cell_map):
    """
    Extract a meaningful title from the table bounds.
    Handles merged cells and multi-row titles.
    """
    top, bottom, left, right = bounds

    # First check for merged cells that span most of the table width
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue

        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue

            cell_key = (r, c)
            if cell_key in merged_cell_map:
                merged_info = merged_cell_map[cell_key]
                # If it spans a significant portion of the table width
                span_ratio = (merged_info["max_col"] - merged_info["min_col"] + 1) / (
                    right - left + 1
                )
                if span_ratio >= 0.5 and is_nonempty(matrix[r][c]):
                    return str(matrix[r][c]["value"]).strip(), r < top

    # Look for a row with a single non-empty cell that might be a title
    for r in range(max(0, top - 2), top + 2):
        if r >= len(matrix):
            continue

        non_empty_cells = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue

            if is_nonempty(matrix[r][c]):
                non_empty_cells.append((r, c))

        if len(non_empty_cells) == 1:
            r_idx, c_idx = non_empty_cells[0]
            return str(matrix[r_idx][c_idx]["value"]).strip(), r_idx < top

    # Combine header cells from the first row
    title_parts = []
    for c in range(left, right + 1):
        if c >= len(matrix[0]):
            continue

        if is_nonempty(matrix[top][c]) and matrix[top][c]["merged_into"] is None:
            title_parts.append(str(matrix[top][c]["value"]).strip())

    if title_parts:
        return " ".join(title_parts), False

    return "Untitled Table", False


def dataframe_to_html(matrix, bounds, merged_cell_map):
    """
    Convert a submatrix to an HTML table with proper handling of merged cells.
    No styling for AI parsing.
    """
    top, bottom, left, right = bounds

    html_lines = []
    html_lines.append("<table>")

    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue

        row_html = ["<tr>"]
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue

            # Skip cells that are part of a merged region but not the top-left cell
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    continue

            cell = matrix[r][c]
            attrs = ""

            # Add rowspan and colspan for merged cells
            if (r, c) in merged_cell_map:
                merged_info = merged_cell_map[(r, c)]
                rowspan = merged_info["rowspan"]
                colspan = merged_info["colspan"]

                # Only add if the merged region is within our table bounds
                effective_rowspan = min(rowspan, bottom - r + 1)
                effective_colspan = min(colspan, right - c + 1)

                if effective_rowspan > 1:
                    attrs += f' rowspan="{effective_rowspan}"'
                if effective_colspan > 1:
                    attrs += f' colspan="{effective_colspan}"'

            # Add cell content
            if is_nonempty(cell):
                cell_content = escape(str(cell["value"]))
            else:
                cell_content = ""

            row_html.append(f"<td{attrs}>{cell_content}</td>")

        row_html.append("</tr>")
        html_lines.append("".join(row_html))

    html_lines.append("</table>")
    return "\n".join(html_lines)


def dataframe_to_markdown(matrix, bounds, merged_cell_map):
    """
    Convert a submatrix to a Markdown table.
    Note: Markdown doesn't support merged cells natively, so we'll handle them specially.
    """
    top, bottom, left, right = bounds

    # First, create a 2D grid to represent the table
    # We'll fill it with cell values and then convert to Markdown
    table_grid = []
    for r in range(top, bottom + 1):
        if r >= len(matrix):
            continue

        row = []
        for c in range(left, right + 1):
            if c >= len(matrix[0]):
                continue

            # For merged cells, we'll only include the content in the top-left cell
            # and use empty strings for the other cells in the merged region
            if matrix[r][c]["merged_into"] is not None:
                merged_r, merged_c = matrix[r][c]["merged_into"]
                if merged_r != r or merged_c != c:
                    row.append("")
                    continue

            cell = matrix[r][c]
            if is_nonempty(cell):
                # Escape pipe characters in Markdown tables
                cell_content = str(cell["value"]).strip().replace("|", "\\|")
            else:
                cell_content = ""

            row.append(cell_content)

        table_grid.append(row)

    if not table_grid:
        return "*Empty table*"

    # Calculate column widths for better formatting
    col_widths = [0] * len(table_grid[0])
    for row in table_grid:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(cell))

    # Generate the Markdown table
    md_lines = []

    # Header row
    header = (
        "| "
        + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(table_grid[0]))
        + " |"
    )
    md_lines.append(header)

    # Separator row
    separator = "| " + " | ".join("-" * max(3, width) for width in col_widths) + " |"
    md_lines.append(separator)

    # Data rows
    for row_idx, row in enumerate(table_grid):
        if row_idx == 0:  # Skip the header row, already added
            continue

        md_row = (
            "| "
            + " | ".join(cell.ljust(col_widths[i]) for i, cell in enumerate(row))
            + " |"
        )
        md_lines.append(md_row)

    return "\n".join(md_lines)


def process_sheet_html(ws):
    """
    Process a worksheet to extract subtables with smart detection.
    Output in HTML format.
    """
    matrix, merged_cell_map = get_sheet_matrix(ws)
    nrows = len(matrix)
    if nrows == 0:
        return "<p>No data found in this sheet.</p>"

    # Use the smart BFS to detect non-overlapping clusters
    blocks = smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1)

    html_fragments = []
    for bounds in blocks:
        top, bottom, left, right = bounds

        # Extract title with improved handling for merged cells
        title, is_external_title = extract_title(matrix, bounds, merged_cell_map)

        html_fragments.append(f"<h3>{escape(title)}</h3>")

        # Adjust table bounds if the title is part of the table
        table_top = (
            top + 1 if not is_external_title and title != "Untitled Table" else top
        )

        if table_top <= bottom:
            table_html = dataframe_to_html(
                matrix, (table_top, bottom, left, right), merged_cell_map
            )
            html_fragments.append(table_html)

        html_fragments.append("<hr>")

    return "\n".join(html_fragments)


def process_sheet_markdown(ws):
    """
    Process a worksheet to extract subtables with smart detection.
    Output in Markdown format.
    """
    matrix, merged_cell_map = get_sheet_matrix(ws)
    nrows = len(matrix)
    if nrows == 0:
        return "*No data found in this sheet.*"

    # Use the smart BFS to detect non-overlapping clusters
    blocks = smart_bfs_components(matrix, merged_cell_map, max_v_gap=2, max_h_gap=1)

    md_fragments = []
    for bounds in blocks:
        top, bottom, left, right = bounds

        # Extract title with improved handling for merged cells
        title, is_external_title = extract_title(matrix, bounds, merged_cell_map)

        md_fragments.append(f"### {title}")
        md_fragments.append("")  # Empty line after title

        # Adjust table bounds if the title is part of the table
        table_top = (
            top + 1 if not is_external_title and title != "Untitled Table" else top
        )

        if table_top <= bottom:
            table_md = dataframe_to_markdown(
                matrix, (table_top, bottom, left, right), merged_cell_map
            )
            md_fragments.append(table_md)
            md_fragments.append("")  # Empty line after table

        md_fragments.append("---")
        md_fragments.append("")  # Empty line after separator

    return "\n".join(md_fragments)


def create_html_from_workbook_bytes(file_bytes: bytes) -> str:
    """
    Process each sheet in the workbook (provided as bytes) to extract subtables.
    Returns HTML output.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error loading workbook: {e}")

    html_lines = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        '    <meta charset="UTF-8">',
        "    <title>Extracted Excel Subtables</title>",
        "</head>",
        "<body>",
        "<h1>Extracted Subtables</h1>",
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_lines.append("<section>")
        html_lines.append(f"<h2>Sheet: {escape(sheet_name)}</h2>")
        sheet_html = process_sheet_html(ws)
        html_lines.append(sheet_html)
        html_lines.append("</section>")

    html_lines.append("</body>")
    html_lines.append("</html>")
    return "\n".join(html_lines)


def create_markdown_from_workbook_bytes(file_bytes: bytes) -> str:
    """
    Process each sheet in the workbook (provided as bytes) to extract subtables.
    Returns Markdown output.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error loading workbook: {e}")

    md_lines = [
        "# Extracted Excel Subtables",
        "",
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"## Sheet: {sheet_name}")
        md_lines.append("")
        sheet_md = process_sheet_markdown(ws)
        md_lines.append(sheet_md)
        md_lines.append("")

    return "\n".join(md_lines)


@app.post("/extract")
async def extract_subtables(
    file: UploadFile = File(...),
    output_format: str = Query(
        "html", regex="^(html|md)$", description="Output format: 'html' or 'md'"
    ),
):
    """
    POST endpoint that accepts an XLSX file upload and returns the extracted subtables
    in either HTML or Markdown format.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        file_bytes = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading the file: {e}")

    if output_format.lower() == "md":
        result = create_markdown_from_workbook_bytes(file_bytes)
        media_type = "text/markdown"
    else:
        result = create_html_from_workbook_bytes(file_bytes)
        media_type = "text/html"

    return Response(content=result, media_type=media_type)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app)

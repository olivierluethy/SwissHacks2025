import argparse
from collections import deque
from openpyxl import load_workbook
from html import escape
import sys

# A list of RGB values that we consider as "default" (non‐header).
# Usually, a cell with no fill or a white background is not marked as header.
DEFAULT_FILL_COLORS = {None, '00000000', 'FFFFFFFF'}

def get_cell_info(cell):
    """
    Return a dictionary with the cell's value and its fill color (RGB string if available).
    We use cell.fill.start_color.rgb if the fill type is set.
    """
    val = cell.value
    color = None
    if cell.fill and cell.fill.fill_type:
        # openpyxl stores RGB in a string like 'FF0000FF' (for blue, for instance).
        color = cell.fill.start_color.rgb
    return {"value": val, "color": color}

def get_sheet_matrix(ws):
    """
    Read the worksheet into a 2D matrix of dictionaries (one per cell)
    preserving both the cell value and its fill color.
    The matrix uses 0-based indexing.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    matrix = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            row_vals.append(get_cell_info(ws.cell(row=r, column=c)))
        matrix.append(row_vals)
    return matrix

def is_nonempty(cell_info):
    """
    A cell is considered nonempty if its value is not None or not just whitespace.
    """
    val = cell_info["value"]
    return val is not None and str(val).strip() != ""

def is_header_cell(cell_info):
    """
    We consider a cell as a header candidate if it has a fill color that is not the default.
    You might adjust this logic if your files use specific colors.
    """
    color = cell_info["color"]
    return color not in DEFAULT_FILL_COLORS

def bfs_components(matrix):
    """
    Use BFS to extract connected components (clusters) of nonempty cells.
    Two cells are connected if they share a side.
    Returns a list of bounding boxes (top, bottom, left, right) with 0-based indices.
    """
    nrows = len(matrix)
    ncols = len(matrix[0]) if nrows else 0
    visited = [[False] * ncols for _ in range(nrows)]
    components = []
    
    for r in range(nrows):
        for c in range(ncols):
            if not visited[r][c] and is_nonempty(matrix[r][c]):
                # Start a BFS from this cell.
                q = deque()
                q.append((r, c))
                visited[r][c] = True
                min_r, max_r = r, r
                min_c, max_c = c, c
                while q:
                    cr, cc = q.popleft()
                    # Update boundaries.
                    min_r = min(min_r, cr)
                    max_r = max(max_r, cr)
                    min_c = min(min_c, cc)
                    max_c = max(max_c, cc)
                    # Check the four neighbors.
                    for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                        nr, nc = cr + dr, cc + dc
                        if 0 <= nr < nrows and 0 <= nc < ncols:
                            if not visited[nr][nc] and is_nonempty(matrix[nr][nc]):
                                visited[nr][nc] = True
                                q.append((nr, nc))
                components.append((min_r, max_r, min_c, max_c))
    return components

def expand_upward_for_headers(matrix, bounds, max_expand=3):
    """
    Given a block defined by bounds (top, bottom, left, right), look upward for up to max_expand rows.
    If in the same horizontal region (left to right) there are cells that are header candidates,
    extend the block upward to include them.
    """
    top, bottom, left, right = bounds
    nrows = len(matrix)
    new_top = top
    # Look upward; allow gaps if the header cell is present.
    for i in range(1, max_expand + 1):
        check_row = top - i
        if check_row < 0:
            break
        # Check cells in the horizontal region.
        header_found = False
        for col in range(left, right + 1):
            cell = matrix[check_row][col]
            if is_header_cell(cell):
                header_found = True
                break
        if header_found:
            new_top = check_row  # absorb this row into our block
        else:
            # If no header cell is found in this row, but the row is not completely empty,
            # you might want to consider absorbing it as well. For now, we require an explicit header color.
            continue
    return (new_top, bottom, left, right)

def dataframe_to_html(matrix, bounds):
    """
    Convert a submatrix (given by bounds: top, bottom, left, right) to an HTML table.
    Empty cells are replaced by a non-breaking space.
    """
    top, bottom, left, right = bounds
    html_lines = []
    html_lines.append('<table border="1" style="border-collapse:collapse;text-align:left;">')
    for r in range(top, bottom + 1):
        row_html = ["<tr>"]
        for c in range(left, right + 1):
            cell = matrix[r][c]
            if is_nonempty(cell):
                cell_content = escape(str(cell["value"]))
            else:
                cell_content = "&nbsp;"
            row_html.append(f"<td>{cell_content}</td>")
        row_html.append("</tr>")
        html_lines.append("".join(row_html))
    html_lines.append("</table>")
    return "\n".join(html_lines)

def process_sheet(ws):
    """
    Process an openpyxl worksheet: build a cell matrix (with color info),
    use BFS to find clusters, expand clusters upward to include header rows
    (by looking at cell fill colors), and produce HTML for each subtable.
    """
    matrix = get_sheet_matrix(ws)
    nrows = len(matrix)
    if nrows == 0:
        return "<p>No data found in this sheet.</p>"
    ncols = len(matrix[0])
    
    # Find clusters of adjacent nonempty cells.
    blocks = bfs_components(matrix)
    
    # Expand each block upward based on header cell fill colors.
    expanded_blocks = []
    for bounds in blocks:
        expanded = expand_upward_for_headers(matrix, bounds, max_expand=3)
        expanded_blocks.append(expanded)
    
    # Optional: sort blocks by their top and left positions.
    expanded_blocks.sort(key=lambda b: (b[0], b[2]))
    
    html_fragments = []
    for bounds in expanded_blocks:
        top, bottom, left, right = bounds
        
        # Check the first row for a single non-empty cell to use as title
        title = "Untitled Subtable"
        skip_first_row = False
        if top <= bottom:  # Ensure there's at least one row
            first_row = matrix[top][left:right + 1]
            non_empty_cells = [cell for cell in first_row if is_nonempty(cell)]
            if len(non_empty_cells) == 1 and all(not is_nonempty(matrix[top][c]) for c in range(left, right + 1) if c != left):
                # First row has exactly one non-empty cell in the leftmost column
                title = str(non_empty_cells[0]["value"]).strip()
                skip_first_row = True
        
        # Set the title
        html_fragments.append(f"<h3>{escape(title)}</h3>")
        
        # Adjust the top boundary if skipping the first row
        table_top = top + 1 if skip_first_row else top
        
        # Convert the block to an HTML table
        if table_top <= bottom:  # Only generate table if there's data left
            table_html = dataframe_to_html(matrix, (table_top, bottom, left, right))
            html_fragments.append(table_html)
        html_fragments.append("<hr>")
    
    return "\n".join(html_fragments)
    """
    Process an openpyxl worksheet: build a cell matrix (with color info),
    use BFS to find clusters, expand clusters upward to include header rows
    (by looking at cell fill colors), and produce HTML for each subtable.
    """
    matrix = get_sheet_matrix(ws)
    nrows = len(matrix)
    if nrows == 0:
        return "<p>No data found in this sheet.</p>"
    ncols = len(matrix[0])
    
    # Find clusters of adjacent nonempty cells.
    blocks = bfs_components(matrix)
    
    # Expand each block upward based on header cell fill colors.
    expanded_blocks = []
    for bounds in blocks:
        expanded = expand_upward_for_headers(matrix, bounds, max_expand=3)
        expanded_blocks.append(expanded)
    
    # Optional: sort blocks by their top and left positions.
    expanded_blocks.sort(key=lambda b: (b[0], b[2]))
    
    html_fragments = []
    for bounds in expanded_blocks:
        top, bottom, left, right = bounds
        
        # Try to form a title from the top rows of the block if header cells are present.
        title_lines = []
        # We take the top 1-2 rows of the block as candidate header lines if they contain header cells.
        header_end = min(top + 2, bottom + 1)
        for r in range(top, header_end):
            row_vals = []
            for c in range(left, right + 1):
                cell = matrix[r][c]
                # If the cell is considered a header cell, include its text.
                if is_header_cell(cell) and is_nonempty(cell):
                    row_vals.append(str(cell["value"]).strip())
            if row_vals:
                title_lines.append(" ".join(row_vals))
        if title_lines:
            title = " | ".join(title_lines)
            html_fragments.append(f"<h3>{escape(title)}</h3>")
        else:
            html_fragments.append("<h3>Untitled Subtable</h3>")
        
        # Convert the block to an HTML table.
        table_html = dataframe_to_html(matrix, bounds)
        html_fragments.append(table_html)
        html_fragments.append("<hr>")
    
    return "\n".join(html_fragments)

def create_html_from_workbook(filepath):
    """
    Load the workbook, process each sheet to extract subtables (using BFS and header cell detection),
    and build a single HTML document.
    """
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print("Error loading workbook:", e)
        sys.exit(1)
    
    html_lines = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        '    <meta charset="UTF-8">',
        "    <title>Extracted Excel Subtables</title>",
        "</head>",
        "<body>",
        "<h1>Extracted Subtables</h1>"
    ]
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_lines.append(f'<section style="margin-bottom: 2em;">')
        html_lines.append(f"<h2>Sheet: {escape(sheet_name)}</h2>")
        sheet_html = process_sheet(ws)
        html_lines.append(sheet_html)
        html_lines.append("</section>")
    
    html_lines.append("</body>")
    html_lines.append("</html>")
    return "\n".join(html_lines)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description="Extract subtables from an Excel file using BFS and header detection by cell colors. "
                    "Handles multiple blocks even if headers are detached by a gap."
    )
    parser.add_argument("input_file", help="Path to the input Excel file (.xlsx)")
    parser.add_argument("output_file", help="Path for the output HTML file")
    args = parser.parse_args()
    
    final_html = create_html_from_workbook(args.input_file)
    with open(args.output_file, "w", encoding="utf-8") as f:
        f.write(final_html)
